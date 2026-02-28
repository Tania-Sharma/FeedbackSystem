from fileinput import filename
import os

from flask import Flask, render_template, redirect, url_for, flash, request
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import random
from openpyxl import load_workbook

from models import db, User, Feedback, PasswordReset, UploadedFile

app = Flask(__name__)
app.config['SECRET_KEY'] = 'college-feedback-secret-key-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///feedback.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

BRANCHES = ['CSE-AI&DS', 'Civil', 'Electrical']

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ========================
# HOME ROUTE
# ========================
@app.route('/')
def home():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            if getattr(current_user, 'sub_role', None) == 'superadmin':
                return redirect(url_for('superadmin_dashboard'))
            else:
                return redirect(url_for('admin_dashboard'))
        elif current_user.role == 'teacher':
            return redirect(url_for('teacher_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))
    return redirect(url_for('login'))

# ========================
# LOGIN / LOGOUT
# ========================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        selected_role = request.form.get('role')
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            if user.role.lower() == selected_role.lower():
                login_user(user)
                return redirect(url_for('home'))
            else:
                flash("You selected wrong login type!", "danger")
        else:
            flash("Invalid username or password!", "danger")
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully!', 'info')
    return redirect(url_for('login'))

# ========================
# FORGOT / RESET PASSWORD
# ========================
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form.get('username')).first()
        if user:
            code = ''.join(random.choices('0123456789', k=6))
            reset = PasswordReset(
                user_id=user.id,
                reset_code=code,
                expires_at=datetime.utcnow() + timedelta(minutes=30)
            )
            db.session.add(reset)
            db.session.commit()
            flash(f'Reset code: {code}', 'info')
            return redirect(url_for('reset_password'))
        flash('Username not found!', 'danger')
    return render_template('forgot_password.html')

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'POST':
        reset = PasswordReset.query.filter_by(
            reset_code=request.form.get('reset_code'),
            is_used=False
        ).first()
        if reset and reset.expires_at > datetime.utcnow():
            user = User.query.get(reset.user_id)
            user.password = generate_password_hash(request.form.get('new_password'))
            reset.is_used = True
            db.session.commit()
            flash('Password reset successful! Login now.', 'success')
            return redirect(url_for('login'))
        flash('Invalid or expired reset code!', 'danger')
    return render_template('reset_password.html')

# ========================
# STUDENT DASHBOARD
# ========================
@app.route('/student')
@login_required
def student_dashboard():
    if current_user.role != 'student':
        return redirect(url_for('home'))
    teachers = User.query.filter_by(role='teacher', branch=current_user.branch).all()
    submitted_feedbacks = Feedback.query.filter_by(student_id=current_user.id).all()
    submitted_ids = [f.teacher_id for f in submitted_feedbacks]
    return render_template('student/dashboard.html', teachers=teachers, submitted_ids=submitted_ids, branches=BRANCHES)

@app.route('/submit_feedback/<int:teacher_id>', methods=['POST'])
@login_required
def submit_feedback(teacher_id):
    if current_user.role != 'student':
        return redirect(url_for('home'))
    existing = Feedback.query.filter_by(student_id=current_user.id, teacher_id=teacher_id, semester=current_user.semester).first()
    if existing:
        flash('Feedback already submitted for this teacher this semester!', 'warning')
    else:
        new_feedback = Feedback(
            student_id=current_user.id,
            teacher_id=teacher_id,
            rating=int(request.form.get('rating')),
            comment=request.form.get('comment'),
            semester=current_user.semester,
            academic_year=current_user.academic_year or '2024-2025'
        )
        db.session.add(new_feedback)
        db.session.commit()
        flash('Feedback submitted successfully!', 'success')
    return redirect(url_for('student_dashboard'))

# ========================
# TEACHER DASHBOARD
# ========================
@app.route('/teacher')
@login_required
def teacher_dashboard():
    if current_user.role != 'teacher':
        return redirect(url_for('home'))
    feedbacks = Feedback.query.filter_by(teacher_id=current_user.id).order_by(Feedback.created_at.desc()).all()
    total = len(feedbacks)
    avg = sum(f.rating for f in feedbacks) / total if total else 0
    excellent = sum(1 for f in feedbacks if f.rating==5)
    good = sum(1 for f in feedbacks if f.rating==4)
    average = sum(1 for f in feedbacks if f.rating==3)
    poor = sum(1 for f in feedbacks if f.rating==2)
    very_poor = sum(1 for f in feedbacks if f.rating==1)
    return render_template('teacher/dashboard.html', feedbacks=feedbacks, total_feedbacks=total, average_rating=round(avg,1),
                           excellent_count=excellent, good_count=good, average_count=average, poor_count=poor, very_poor_count=very_poor)

# ========================
# BRANCH ADMIN DASHBOARD
# ========================
@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'branchadmin':
        return redirect(url_for('home'))
    selected_branch = current_user.branch
    selected_semester = request.args.get('semester', None, type=int)
    selected_teacher = request.args.get('teacher', None, type=int)

    students_query = User.query.filter_by(role='student', branch=selected_branch)
    if selected_semester:
        students_query = students_query.filter_by(semester=selected_semester)
    students = students_query.all()
    teachers = User.query.filter_by(role='teacher', branch=selected_branch).all()

    feedback_query = Feedback.query.join(User, Feedback.student_id==User.id).filter(User.branch==selected_branch)
    if selected_semester:
        feedback_query = feedback_query.filter(Feedback.semester==selected_semester)
    if selected_teacher:
        feedback_query = feedback_query.filter(Feedback.teacher_id==selected_teacher)
    feedbacks = feedback_query.all()

    total = len(feedbacks)
    avg = sum(f.rating for f in feedbacks)/total if total else 0
    excellent = sum(1 for f in feedbacks if f.rating==5)
    good = sum(1 for f in feedbacks if f.rating==4)
    average = sum(1 for f in feedbacks if f.rating==3)
    poor = sum(1 for f in feedbacks if f.rating==2)
    very_poor = sum(1 for f in feedbacks if f.rating==1)

    return render_template('admin/dashboard.html', students=students, teachers=teachers, branches=BRANCHES,
                           selected_branch=selected_branch, selected_semester=selected_semester,
                           selected_teacher=selected_teacher, total_feedbacks=total,
                           average_rating=round(avg,1), excellent_count=excellent, good_count=good,
                           average_count=average, poor_count=poor, very_poor_count=very_poor)

# ========================
# SUPERADMIN DASHBOARD
# ========================
@app.route('/superadmin_dashboard')
@login_required
def superadmin_dashboard():
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    admins = User.query.filter_by(role='admin').all()
    students_files = UploadedFile.query.filter_by(file_type='student').all()
    teachers_files = UploadedFile.query.filter_by(file_type='teacher').all()
    return render_template(
    'admin/superadmin_dashboard.html',
    admins=admins,
    branches=BRANCHES,
    all_users=User.query.all(),
    total_students=User.query.filter_by(role='student').count(),
    total_teachers=User.query.filter_by(role='teacher').count(),
    total_feedbacks=Feedback.query.count(),
    current_semester="6",  # ya jo bhi hai
    students_files=students_files,
    teachers_files=teachers_files
)

@app.route('/add_admin', methods=['POST'])
@login_required
def add_admin():
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    username = request.form['username']
    password = request.form['password']
    sub_role = request.form['sub_role']
    branch = request.form.get('branch') if sub_role=='branchadmin' else None
    hashed_pw = generate_password_hash(password)
    new_admin = User(username=username, password=hashed_pw, role='admin', sub_role=sub_role, branch=branch)
    db.session.add(new_admin)
    db.session.commit()
    flash('New admin added successfully!', 'success')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/delete_admin/<int:admin_id>')
@login_required
def delete_admin(admin_id):
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    admin = User.query.get(admin_id)
    if admin and admin.sub_role != 'superadmin':
        db.session.delete(admin)
        db.session.commit()
        flash('Admin deleted successfully!', 'success')
    else:
        flash('Cannot delete superadmin!', 'danger')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/change_admin_password/<int:admin_id>', methods=['POST'])
@login_required
def change_admin_password(admin_id):
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    new_password = request.form['new_password']
    admin = User.query.get(admin_id)
    admin.password = generate_password_hash(new_password)
    db.session.commit()
    flash('Password changed successfully!', 'success')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/change_admin_branch/<int:admin_id>', methods=['POST'])
@login_required
def change_admin_branch(admin_id):
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    new_branch = request.form['new_branch']
    admin = User.query.get(admin_id)
    admin.branch = new_branch
    db.session.commit()
    flash('Branch updated successfully!', 'success')
    return redirect(url_for('superadmin_dashboard'))

# ========================
# EXCEL UPLOAD / DELETE
# ========================
@app.route('/upload_students', methods=['POST'])
@login_required
def upload_students():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    
    branch = request.form.get('branch')
    semester = request.form.get('semester')

    if not semester or str(semester).strip() == '':
        semester = 1
    else:
        semester = int(semester)
    
    if 'file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(url_for('admin_dashboard', branch=branch, semester=semester))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('admin_dashboard', branch=branch, semester=semester))
    
    try:
        # ✅ Excel read for importing students
        wb = load_workbook(file)
        ws = wb.active
        
        imported = 0
        skipped = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                continue
            
            username = row[0]
            if not username:
                continue
            
            if User.query.filter_by(username=str(username)).first():
                skipped += 1
                continue
            
            user = User(
                username=str(username),
                password=generate_password_hash(str(row[1]) if row[1] else '123456'),
                role='student',
                name=str(row[2]) if row[2] else '',
                email=str(row[3]) if row[3] else '',
                branch=branch,
                semester=semester,
                academic_year=str(row[4]) if len(row) > 4 and row[4] else '2024-2025',
                roll_number=str(row[5]) if len(row) > 5 and row[5] else '',
                imported_file_id=uploaded_file.id   # ⭐ ADD THIS
        )
            db.session.add(user)
            imported += 1
        
        db.session.commit()
        flash(f'Successfully imported {imported} students! ({skipped} skipped)', 'success')

        import os

        file.seek(0)

        upload_folder = os.path.join(os.getcwd(), 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)

        filename = file.filename
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)

# Create UploadedFile entry
        uploaded_file = UploadedFile(
            filename=filename,
            file_type='student',
            uploaded_by=current_user.id
        )

        db.session.add(uploaded_file)
        db.session.commit()
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('admin_dashboard', branch=branch, semester=semester))

@app.route('/upload_teachers', methods=['POST'])
@login_required
def upload_teachers():
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    if 'file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        imported = 0
        skipped = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                continue
            
            username = row[0]
            if not username:
                continue
            
            if User.query.filter_by(username=str(username)).first():
                skipped += 1
                continue
            
            user = User(
                username=str(username),
                password=generate_password_hash(str(row[1]) if row[1] else '123456'),
                role='teacher',
                name=str(row[2]) if row[2] else '',
                email=str(row[3]) if row[3] else '',
                subject=str(row[4]) if len(row) > 4 else ''
            )
            db.session.add(user)
            imported += 1
        
        db.session.commit()
        flash(f'Successfully imported {imported} teachers! ({skipped} skipped)', 'success')
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/upload_admins', methods=['POST'])
@login_required
def upload_admins():
    if current_user.role != 'admin' or getattr(current_user, 'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))

    if 'file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(url_for('superadmin_dashboard'))

    file = request.files['file']
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('superadmin_dashboard'))

    try:
        wb = load_workbook(file)
        ws = wb.active
        imported = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            username = row[0]
            if not username or User.query.filter_by(username=username).first():
                skipped += 1
                continue
            user = User(
                username=username,
                password=generate_password_hash(row[1] if row[1] else 'admin123'),
                role='admin',
                sub_role=row[2] if len(row) > 2 else 'branchadmin',
                branch=row[3] if len(row) > 3 else None,
                name=row[4] if len(row) > 4 else '',
                email=row[5] if len(row) > 5 else ''
            )
            db.session.add(user)
            imported += 1
        db.session.commit()
        flash(f'Successfully imported {imported} admins! ({skipped} skipped)', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/delete_file/<int:file_id>')
@login_required
def delete_file(file_id):
    file = UploadedFile.query.get(file_id)

    if not file:
        flash("File not found!", "danger")
        return redirect(url_for('admin_dashboard'))

    # Delete related students
    User.query.filter_by(imported_file_id=file.id).delete()

    # Delete physical file
    upload_folder = os.path.join(os.getcwd(), 'uploads')
    file_path = os.path.join(upload_folder, file.filename)

    if os.path.exists(file_path):
        os.remove(file_path)

    # Delete file record
    db.session.delete(file)
    db.session.commit()

    flash("File and related students deleted permanently!", "success")
    return redirect(url_for('admin_dashboard'))

    flash("File and related students deleted permanently!", "success")
    return redirect(url_for('admin_dashboard'))

# ========================
# VIEW / EDIT / DELETE STUDENTS & TEACHERS
# ========================
@app.route('/view_students')
@login_required
def view_students():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    students = User.query.filter_by(role='student', branch=current_user.branch).all() \
               if getattr(current_user, 'sub_role', None)=='branchadmin' else User.query.filter_by(role='student').all()
    return render_template('admin/view_students.html', users=students)

@app.route('/view_teachers')
@login_required
def view_teachers():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    teachers = User.query.filter_by(role='teacher').all()
    return render_template('admin/view_teachers.html', users=teachers)

@app.route('/edit_user/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    user = User.query.get(user_id)
    if not user:
        flash("User not found!", "danger")
        return redirect(request.referrer)
    
    # Update fields
    user.name = request.form.get('name', user.name)
    user.email = request.form.get('email', user.email)
    if user.role=='student':
        user.branch = request.form.get('branch', user.branch)
        user.semester = int(request.form.get('semester', user.semester))
        user.academic_year = request.form.get('academic_year', user.academic_year)
        user.roll_number = request.form.get('roll_number', user.roll_number)
    elif user.role=='teacher':
        user.subject = request.form.get('subject', user.subject)
    
    db.session.commit()
    flash(f"{user.role.title()} '{user.username}' updated successfully!", "success")
    return redirect(request.referrer)

@app.route('/delete_user/<int:user_id>')
@login_required
def delete_user(user_id):
    user = User.query.get(user_id)

    if not user:
        flash("User not found!", "danger")
        return redirect(url_for('admin_dashboard'))

    db.session.delete(user)
    db.session.commit()

    flash("User deleted permanently!", "success")
    return redirect(url_for('admin_dashboard'))
@app.route('/view_uploaded_files')
@login_required
def view_uploaded_files():
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    students_files = UploadedFile.query.filter_by(file_type='student').all()
    teachers_files = UploadedFile.query.filter_by(file_type='teacher').all()
    admins_files = UploadedFile.query.filter_by(file_type='admin').all()

    return render_template('admin/view_data_files.html',
                           students_files=students_files,
                           teachers_files=teachers_files,
                           admins_files=admins_files)

from flask import send_from_directory

@app.route('/open_file/<int:file_id>')
@login_required
def open_file(file_id):
    file = UploadedFile.query.get(file_id)

    upload_folder = os.path.join(os.getcwd(), 'uploads')

    return send_from_directory(upload_folder, file.filename)

@app.route('/update_file/<int:file_id>', methods=['POST'])
@login_required
def update_file(file_id):
    file = UploadedFile.query.get(file_id)
    if not file:
        flash("File not found!", "danger")
        return redirect(request.referrer)

    file.content = request.form.get('content')
    db.session.commit()

    flash("File updated successfully!", "success")
    return redirect(url_for('view_uploaded_files'))

# ========================
# INITIALIZE DATABASE
# ========================
def create_default_admin():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(username='superadmin').first():
            superadmin = User(username='superadmin', password=generate_password_hash('admin123'),
                              email='superadmin@college.com', role='admin', sub_role='superadmin',
                              name='Super Admin', branch=None)
            db.session.add(superadmin)
        for branch in BRANCHES:
            uname = branch.lower().replace('&','').replace('-','').replace(' ','')+'admin'
            if not User.query.filter_by(username=uname).first():
                admin = User(username=uname, password=generate_password_hash('admin123'),
                             email=f'{uname}@college.com', role='admin', sub_role='branchadmin',
                             name=f'{branch} Admin', branch=branch)
                db.session.add(admin)
        db.session.commit()
        print("✅ Database initialized successfully!")
        print("\n📋 Default Login Credentials:")
        print("-"*50)
        print("| Role          | Username          | Password |")
        print("-"*50)
        print("| Super Admin   | superadmin        | admin123 |")
        for branch in BRANCHES:
            uname = branch.lower().replace('&','').replace('-','').replace(' ','')+'admin'
            print(f"| {branch} Admin | {uname:<15} | admin123 |")
        print("-"*50)

if __name__ == '__main__':
    create_default_admin()
    app.run(debug=True)