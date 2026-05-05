import os
import io
import random
from datetime import datetime, timedelta

from flask import (
    Flask, render_template, redirect, url_for,
    flash, request, send_file, jsonify, session
)

from flask_login import (
    LoginManager, login_user, logout_user,
    login_required, current_user
)

from werkzeug.security import generate_password_hash, check_password_hash

from flask_mail import Mail, Message

import pandas as pd
from openpyxl import load_workbook

from models import (
    db, User, Feedback, PasswordReset,
    UploadedFile, TeacherSubject,
    FeedbackWindow, AdminSettings
)

app = Flask(__name__)
app.secret_key = "anything_random"


def flexible_header_match(header_list, targets):
    """Excel headers mein se keywords dhoondhne ke liye dynamic logic"""
    
    # targets ko bhi clean kar lo
    targets = [
        str(t).strip().lower().replace(' ', '').replace('_', '').replace('-', '').replace('.', '')
        for t in targets
    ]

    for i, h in enumerate(header_list):
        if h:
            clean_h = str(h).strip().lower().replace(' ', '').replace('_', '').replace('-', '').replace('.', '')
            
            for target in targets:
                if target in clean_h:
                    return i   # match mila → index return

    return None  # kuch match nahi mila

def normalize_branch(branch):
    """Branch name ko clean karne ke liye helper"""
    if not branch:
        return ""
    branch = str(branch).strip().lower()
    branch = branch.replace("-", "").replace("(", "").replace(")", "").replace(" ", "")
    return branch



def normalize_header(header):
    if not header:
        return ""
    return str(header).strip().lower().replace(" ", "").replace("_", "").replace("-", "")


def parse_semester(value):
    if value is None:
        return None

    value = str(value).strip()

    if value == "":
        return None

    if value.endswith(".0"):
        value = value[:-2]

    digits = ''.join(ch for ch in value if ch.isdigit())
    if digits:
        return int(digits)

    return None
app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), 'uploads')
app.config['SECRET_KEY'] = 'college-feedback-secret-key-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///feedback.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'your_email@gmail.com'
app.config['MAIL_PASSWORD'] = 'your_app_password'
app.config['MAIL_DEFAULT_SENDER'] = 'your_email@gmail.com'
mail = Mail(app)

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///feedback.db?timeout=30'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

def get_all_branches():
    branches = db.session.query(User.branch).filter(User.branch != None).distinct().all()
    return [b[0] for b in branches if b[0]]



@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

from flask import send_from_directory
@app.route('/open_file/<int:file_id>')
@login_required
def open_file(file_id):
    file = UploadedFile.query.get(file_id)

    if not file:
        flash("File not found in database!", "danger")
        return redirect(url_for('superadmin_dashboard'))

    upload_folder = os.path.join(app.root_path, 'uploads')
    file_path = os.path.join(upload_folder, file.filename)

    if not os.path.exists(file_path):
        flash(f"Physical file missing: {file.filename}", "danger")
        return redirect(url_for('superadmin_dashboard'))

    return send_from_directory(upload_folder, file.filename, as_attachment=True)

# ========================
# HOME ROUTE
# ========================
@app.route('/')
def home():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            if getattr(current_user, 'sub_role', None) == 'superadmin':
                return render_template('admin/superadmin_dashboard.html',
                                       students_files=UploadedFile.query.filter_by(file_type='student').all(),
                                       teachers_files=UploadedFile.query.filter_by(file_type='teacher').all(),
                                       admins_files=UploadedFile.query.filter_by(file_type='admin').all())
            else:
                return redirect(url_for('admin_dashboard'))
        elif current_user.role == 'teacher':
            return redirect(url_for('teacher_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))
    return redirect(url_for('login'))

# ========================
# LOGIN / LOGOUT
# ========================

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':
        print("LOGIN ROUTE HIT")

        # =========================
        # GET FORM DATA
        # =========================
        username = request.form.get('username')
        password = request.form.get('password')
        selected_role = request.form.get('role')

        # =========================
        # SAFE CLEANING
        # =========================
        if not username or not password or not selected_role:
            flash("All fields are required!", "danger")
            return redirect(url_for('login'))

        username = username.strip()

        # =========================
        # FETCH USER FROM DB
        # =========================
        user = User.query.filter_by(username=username).first()

        print("USER FOUND:", user)

        # =========================
        # USER VALIDATION
        # =========================
        if not user:
            flash("User not found!", "danger")
            return redirect(url_for('login'))

        # =========================
        # PASSWORD CHECK
        # =========================
        if not check_password_hash(user.password, password):
            flash("Invalid username or password!", "danger")
            return redirect(url_for('login'))

        # =========================
        # ROLE CHECK (TEACHER / STUDENT / ADMIN)
        # =========================
        if user.role and user.role.strip().lower() != selected_role.strip().lower():
            flash("Wrong login type selected!", "danger")
            return redirect(url_for('login'))

        # =========================
        # LOGIN SUCCESS
        # =========================
        login_user(user)

        session['user_id'] = user.id
        session['role'] = user.role

        print("LOGIN SUCCESS:", user.role)

        # =========================
        # ROLE BASED REDIRECT
        # =========================
        role = user.role.strip().lower()

        if role == "teacher":
           return redirect(url_for('teacher_dashboard'))

        elif role == "student":
            return redirect(url_for('student_dashboard'))

        elif role == "admin":
            return redirect(url_for('admin_dashboard'))

        else:
            return redirect(url_for('home'))

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully!', 'info')
    return redirect(url_for('login'))
@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    if current_user.role not in ['teacher', 'admin']:
        return redirect(url_for('home'))

    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    if not current_password or not new_password or not confirm_password:
        flash('All fields are required!', 'danger')
        return redirect(request.referrer or url_for('home'))

    if not check_password_hash(current_user.password, current_password):
        flash('Current password is incorrect!', 'danger')
        return redirect(request.referrer or url_for('home'))

    if new_password != confirm_password:
        flash('New password and confirm password do not match!', 'danger')
        return redirect(request.referrer or url_for('home'))

    if len(new_password) < 6:
        flash('New password must be at least 6 characters long!', 'danger')
        return redirect(request.referrer or url_for('home'))

    try:
        current_user.password = generate_password_hash(new_password)
        db.session.commit()

        if current_user.imported_file_id:
            uploaded_file = db.session.get(UploadedFile, current_user.imported_file_id)

            if uploaded_file:
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)

                if os.path.exists(file_path):
                    df = pd.read_excel(file_path)

                    username_col = None
                    password_col = None

                    for col in df.columns:
                        col_lower = str(col).strip().lower()
                        if col_lower == 'username':
                            username_col = col
                        if col_lower == 'password':
                            password_col = col

                    if username_col and password_col:
                        df.loc[
                            df[username_col].astype(str).str.strip() == str(current_user.username).strip(),
                            password_col
                        ] = new_password
                        df.to_excel(file_path, index=False)

        flash('Password changed successfully!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error updating password: {str(e)}', 'danger')

    return redirect(request.referrer or url_for('home'))
# ========================
# FORGOT / RESET PASSWORD
# ========================
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    role = request.args.get('role')

    if role not in ['teacher', 'admin']:
        flash('Forgot password is only available for teacher and admin accounts.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        username = request.form.get('username')
        user = User.query.filter_by(username=username).first()

        if not user:
            flash('Username not found!', 'danger')
            return render_template('forgot_password.html', role=role)

        if user.role not in ['teacher', 'admin']:
            flash('Forgot password is only available for teacher and admin accounts.', 'danger')
            return redirect(url_for('login'))

        if not user.email:
            flash('No email registered for this account!', 'danger')
            return render_template('forgot_password.html', role=role)

        code = ''.join(random.choices('0123456789', k=6))

        reset = PasswordReset(
            user_id=user.id,
            reset_code=code,
            expires_at=datetime.utcnow() + timedelta(minutes=10),
            is_used=False
        )
        db.session.add(reset)
        db.session.commit()

        try:
            msg = Message(
                subject='Password Reset Code',
                recipients=[user.email]
            )
            msg.body = f"""
Hello {user.name or user.username},

Your password reset code is: {code}

This code will expire in 10 minutes.

If you did not request this, please ignore this email.
"""
            mail.send(msg)

            flash('Reset code sent to your registered email!', 'success')
            return redirect(url_for('reset_password'))

        except Exception as e:
            flash(f'Error sending email: {str(e)}', 'danger')
            return render_template('forgot_password.html', role=role)

    return render_template('forgot_password.html', role=role)
@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'POST':
        reset = PasswordReset.query.filter_by(
            reset_code=request.form.get('reset_code'),
            is_used=False
        ).first()
        if reset and reset.expires_at > datetime.utcnow():
            user = User.query.get(reset.user_id)
            user.password = generate_password_hash(request.form.get('new_password'))
            reset.is_used = True
            db.session.commit()
            flash('Password reset successful! Login now.', 'success')
            return redirect(url_for('login'))
        flash('Invalid or expired reset code!', 'danger')
    return render_template('reset_password.html')
# ========================
# STUDENT DASHBOARD
# ========================
@app.route('/student_dashboard')
@login_required
def student_dashboard():

    #if 'branch' not in session or 'semester' not in session:
        #return redirect(url_for('select_details'))

    # ========================
    # 1. Feedback Window Check
    # ========================
    feedback_window = FeedbackWindow.query.filter_by(is_active=True).first()
    is_feedback_open = False

    if feedback_window:
        now = datetime.now()
        if feedback_window.start_time <= now <= feedback_window.end_time:
            is_feedback_open = True

    # ========================
    # 2. Session-based selection (IMPORTANT FIX)
    # ========================
    
    selected_branch = current_user.branch
    selected_sem = current_user.semester

    teachers = []

    # ========================
    # 3. Fetch teachers only if valid selection exists
    # ========================
    if selected_branch and selected_sem:

        print("SELECTED:", selected_branch, selected_sem)

        all_teachers = db.session.query(
            TeacherSubject.id.label('mapping_id'),
            TeacherSubject.subject,
            TeacherSubject.subject_code,
            TeacherSubject.branch,  # 👈 ADD THIS
            TeacherSubject.semester,
            User.name.label('teacher_name'),
            User.id.label('teacher_id')
        ).join(User, TeacherSubject.teacher_id == User.id).all()

        # ✅ FILTER IN PYTHON
        teachers = [
            t for t in all_teachers
            if normalize_branch(t.branch) == normalize_branch(selected_branch)
            and t.semester == selected_sem
        ]

        print("TEACHERS:", teachers)
       
    # ========================
    # 4. Render page safely
    # ========================
    return render_template(
        'student/dashboard.html',
        teachers=teachers,
        is_feedback_open=is_feedback_open,
        branches=get_all_branches(),
        selected_branch=selected_branch,
        selected_sem=selected_sem
    )
# Feedback Form for a Teacher
# -----------------------
@app.route('/feedback/<int:teacher_id>')
@login_required
def feedback_form(teacher_id):
    if current_user.role != 'student':
        return redirect(url_for('home'))

    # Feedback window hard check
    feedback_window = FeedbackWindow.query.filter_by(is_active=True).order_by(FeedbackWindow.id.desc()).first()

    if not feedback_window:
        flash("Feedback window is not set by admin yet.", "danger")
        return redirect(url_for('student_dashboard'))

    now = datetime.now()
    if not (feedback_window.start_time <= now <= feedback_window.end_time):
        flash("Feedback submission is currently closed.", "danger")
        return redirect(url_for('student_dashboard'))

    teacher = User.query.get_or_404(teacher_id)

    # GET request se values lo
    subject = request.args.get('subject', '').strip()
    subject_code = request.args.get('subject_code', '').strip()
    selected_branch = request.args.get('branch', '').strip()
    selected_semester = request.args.get('semester', type=int)

    # Basic safety check
    if not subject or not selected_branch or not selected_semester:
        flash("Please select branch, semester and subject properly!", "danger")
        return redirect(url_for('student_dashboard'))

    # Current logged-in student ka actual branch + semester
    actual_branch = current_user.branch
    actual_branch_display = current_user.branch if current_user.branch else ''
    actual_semester = current_user.semester

    # Verify student can only access own branch + semester
    branch_ok = (selected_branch == actual_branch)
    semester_ok = (selected_semester == actual_semester)

    if not branch_ok or not semester_ok:
        if not branch_ok and not semester_ok:
            flash(
                f"You entered wrong branch and semester! Your branch is {actual_branch_display} and your semester is {actual_semester}.",
                "danger"
            )
        elif not branch_ok:
            flash(
                f"You entered wrong branch! Your branch is {actual_branch_display}.",
                "danger"
            )
        elif not semester_ok:
            flash(
                f"You entered wrong semester! Your semester is {actual_semester}.",
                "danger"
            )

        return redirect(
            url_for(
                'student_dashboard',
                branch=actual_branch_display,
                semester=actual_semester
            )
        )

    # Duplicate feedback check
    existing = Feedback.query.filter_by(
        student_id=current_user.id,
        teacher_id=teacher_id,
        subject=subject,
        semester=selected_semester
    ).first()

    if existing:
        flash(f"You have already submitted feedback for {subject} with this teacher.", "warning")
        return redirect(
            url_for(
                'student_dashboard',
                branch=actual_branch_display,
                semester=actual_semester
            )
        )

    return render_template(
        'student/feedback.html',
        teacher=teacher,
        subject=subject,
        subject_code=subject_code,
        branch=actual_branch_display,
        semester=actual_semester
    )
# -----------------------
# Submit Feedback
# -----------------------
@app.route('/submit_feedback/<int:teacher_id>', methods=['POST'])
@login_required
def submit_feedback(teacher_id):
    if current_user.role != 'student':
        return redirect(url_for('home'))

    feedback_window = FeedbackWindow.query.filter_by(is_active=True).order_by(FeedbackWindow.id.desc()).first()

    if not feedback_window:
        flash("Feedback window is not set by admin yet.", "danger")
        return redirect(url_for('student_dashboard'))

    now = datetime.now()
    if not (feedback_window.start_time <= now <= feedback_window.end_time):
        flash("Feedback submission time is over.", "danger")
        return redirect(url_for('student_dashboard'))

    subject = request.form.get("subject")
    if subject:
        subject = subject.strip()
    subject_code = request.form.get("subject_code")
    if subject_code:
        subject_code = subject_code.strip()

    branch = request.form.get("branch")
    if branch:
        branch = branch.strip()

    semester = request.form.get("semester", type=int)

    if not subject:
        flash("Subject is required!", "danger")
        return redirect(url_for('student_dashboard'))
    

    if not branch or not semester:
        flash("Branch and semester are required!", "danger")
        return redirect(url_for('student_dashboard'))

    # ✅ VERIFY
    actual_branch = current_user.branch
    actual_semester = current_user.semester

    if actual_branch != branch or actual_semester != semester:
        flash(f"Wrong selection! Your branch is {current_user.branch} and semester is {actual_semester}", "danger")
        return redirect(url_for('student_dashboard'))

    # ✅ DUPLICATE CHECK
    existing = Feedback.query.filter_by(
        student_id=current_user.id,
        teacher_id=teacher_id,
        subject=subject,
        semester=semester
    ).first()

    if existing:
        flash('Feedback already submitted!', 'warning')
        return redirect(url_for('student_dashboard', branch=branch, semester=semester))

    # ✅ RATINGS
    ratings = []
    for i in range(1, 13):
        r = request.form.get(f"q{i}")
        if not r:
            flash("Please answer all questions!", "danger")
            return redirect(url_for(
                'feedback_form',
                teacher_id=teacher_id,
                subject=subject,
                branch=branch,
                semester=semester
            ))
        ratings.append(int(r))

    total = sum(ratings)

    # ✅ SAVE
    new_feedback = Feedback(
        student_id=current_user.id,
        teacher_id=teacher_id,
        subject=subject,
        subject_code=subject_code,
        q1=ratings[0],
        q2=ratings[1],
        q3=ratings[2],
        q4=ratings[3],
        q5=ratings[4],
        q6=ratings[5],
        q7=ratings[6],
        q8=ratings[7],
        q9=ratings[8],
        q10=ratings[9],
        q11=ratings[10],
        q12=ratings[11],
        total=total,
        semester=semester,
        academic_year=current_user.academic_year
    )

    db.session.add(new_feedback)
    db.session.commit()

    flash('Feedback submitted successfully!', 'success')
    return redirect(url_for('student_dashboard', branch=branch, semester=semester))


# ========================
# TEACHER DASHBOARD
# ========================
@app.route('/teacher_dashboard')
@login_required
def teacher_dashboard():

    if current_user.role != 'teacher':
        return redirect(url_for('home'))

    # Get filter values
    # Fetch all subjects taught by this teacher
    teacher_subjects = TeacherSubject.query.filter_by(teacher_id=current_user.id).all()
    branch = request.args.get('branch')
    branches=get_all_branches(),
    semester = request.args.get('semester')
    subject = request.args.get('subject')
    selected_subject = subject

    # Base query
    query = Feedback.query.filter_by(teacher_id=current_user.id)

    # Apply filters if selected
    if branch:
        feedbacks = [
            f for f in query.all()
            if normalize_branch(f.student.branch) == normalize_branch(branch)
        ]

    if semester:
        query = query.filter(Feedback.semester == int(semester))
    if subject:
        query = query.filter(Feedback.subject.ilike(f"%{subject}%"))

    feedbacks = query.order_by(Feedback.created_at.desc()).all()

    total = len(feedbacks)

    avg = sum(f.total for f in feedbacks) / (total * 12) if total else 0

    very_poor = sum(1 for f in feedbacks if f.total < 20)
    poor = sum(1 for f in feedbacks if 20 <= f.total < 30)
    good = sum(1 for f in feedbacks if 30 <= f.total < 40)
    very_good = sum(1 for f in feedbacks if 40 <= f.total < 50)
    excellent = sum(1 for f in feedbacks if f.total >= 50)

    return render_template(
    'teacher/teacher_dashboard.html',
    feedbacks=feedbacks,
    total_feedbacks=total,
    average_rating=round(avg,1),
    excellent_count=excellent,
    very_good_count=very_good,
    good_count=good,
    poor_count=poor,
    very_poor_count=very_poor,
    branches=get_all_branches(),
    teacher_subjects=teacher_subjects,
    selected_subject=selected_subject
)
@app.route('/teacher/export')
@login_required
def export_teacher_feedback():
    if current_user.role != 'teacher':
        return redirect(url_for('home'))

    # Get filter values
    branch = request.args.get('branch')
    branch = normalize_branch(branch)
    semester = request.args.get('semester')
    subject = request.args.get('subject')

    # Base query
    query = Feedback.query.filter_by(teacher_id=current_user.id)

    # Apply filters
    if branch:
        query = query.join(User, Feedback.student_id == User.id).filter(User.branch == branch)

    if semester:
        query = query.filter(Feedback.semester == int(semester))

    if subject:
        query = query.filter(Feedback.subject.ilike(f"%{subject}%"))

    # FIRST feedbacks banao
    feedbacks = query.order_by(Feedback.created_at.desc()).all()

    # THEN check karo
    if not feedbacks:
        flash('No feedback data available to export!', 'warning')
        return redirect(url_for('teacher_dashboard'))

    # Prepare data for Excel
    data = []
    for f in feedbacks:
        student = User.query.get(f.student_id)

        data.append({
            'Name': student.name if student and student.name else student.username if student else '',
            'Roll No': student.roll_number if student else '',
            'Branch': student.branch if student else '',
            'Semester': f.semester,
            'Subject': f.subject,
            'Subject Code': f.subject_code if f.subject_code else 'N/A',
            'Timestamp': f.created_at.strftime('%d-%m-%Y %H:%M') if f.created_at else '',
            'Q1': f.q1,
            'Q2': f.q2,
            'Q3': f.q3,
            'Q4': f.q4,
            'Q5': f.q5,
            'Q6': f.q6,
            'Q7': f.q7,
            'Q8': f.q8,
            'Q9': f.q9,
            'Q10': f.q10,
            'Q11': f.q11,
            'Q12': f.q12,
            'Total': f.total,
            'Average': round(f.total / 12, 2) if f.total else 0
        })

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Feedbacks')

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='teacher_feedback.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
# ========================
# BRANCH ADMIN DASHBOARD
# ========================
@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    # Filters
    feedback_branch = request.args.get('branch')
    feedback_semester = request.args.get('semester', type=int)
    feedback_subject_code = request.args.get('subject_code')
    feedback_teacher = request.args.get('teacher')

    # Students list
    students_query = User.query.filter_by(role='student')
    if feedback_branch:
        students_query = students_query.filter_by(branch=feedback_branch)
    if feedback_semester:
        students_query = students_query.filter_by(semester=feedback_semester)
    students = students_query.all()

    # Teachers list
    teachers = User.query.filter_by(role='teacher').all()
    print("TEACHERS LIST:", teachers)

    # Subject dropdown data
    subjects_query = TeacherSubject.query
    if feedback_branch:
        subjects_query = subjects_query.filter_by(branch=normalize_branch(feedback_branch))
    if feedback_semester:
        subjects_query = subjects_query.filter_by(semester=feedback_semester)

    subject_rows = subjects_query.all()

    subjects = sorted([
        {
            "name": s.subject,
            "code": s.subject_code
        }
        for s in subject_rows if s.subject_code
    ], key=lambda x: x["code"])

    # Stats query
    feedback_query = Feedback.query.join(User, Feedback.student_id == User.id)
    if feedback_branch:
        feedbacks_stats = [
            f for f in feedback_query.all()
            if normalize_branch(f.student.branch) == normalize_branch(feedback_branch)
        ]
    if feedback_semester:
        feedback_query = feedback_query.filter(Feedback.semester == feedback_semester)
    if feedback_teacher:
        feedback_query = feedback_query.filter(Feedback.teacher_id == feedback_teacher)
    if feedback_subject_code:
        feedback_query = feedback_query.filter(Feedback.subject_code == feedback_subject_code)

    feedbacks_stats = feedback_query.all()

    total = len(feedbacks_stats)
    avg = sum(f.total for f in feedbacks_stats) / (total * 12) if total else 0
    excellent = sum(1 for f in feedbacks_stats if f.total >= 50)
    good = sum(1 for f in feedbacks_stats if 40 <= f.total < 50)
    average = sum(1 for f in feedbacks_stats if 30 <= f.total < 40)
    poor = sum(1 for f in feedbacks_stats if 20 <= f.total < 30)
    very_poor = sum(1 for f in feedbacks_stats if f.total < 20)

    # Files for Upload/View Data tabs
    students_files = UploadedFile.query.filter_by(file_type='student').all()
    teachers_files = UploadedFile.query.filter_by(file_type='teacher').all()

    # Feedback tab table query
    feedbacks_tab_query = Feedback.query.join(User, Feedback.student_id == User.id)
    if feedback_branch:
        feedbacks_tab_query = feedbacks_tab_query.filter(User.branch == feedback_branch)
    if feedback_semester:
        feedbacks_tab_query = feedbacks_tab_query.filter(Feedback.semester == feedback_semester)
    if feedback_subject_code:
        feedbacks_tab_query = feedbacks_tab_query.filter(Feedback.subject_code == feedback_subject_code)
    if feedback_teacher:
        feedbacks_tab_query = feedbacks_tab_query.filter(Feedback.teacher_id == feedback_teacher)

    feedbacks = feedbacks_tab_query.order_by(Feedback.created_at.desc()).all()

    print("TOTAL FEEDBACKS IN DB:", Feedback.query.count())
    print("FILTERED FEEDBACKS:", len(feedbacks))
    print("BRANCHES:", get_all_branches())
    print("TEACHERS:", teachers)
    current_feedback_window = FeedbackWindow.query.filter_by(is_active=True).order_by(FeedbackWindow.id.desc()).first()

    return render_template(
        'admin/superadmin_dashboard.html',
        students=students,
        teachers=teachers,
        subjects=subjects,
        branches=get_all_branches(),
        total_feedbacks=total,
        average_rating=round(avg, 1),
        excellent_count=excellent,
        good_count=good,
        average_count=average,
        poor_count=poor,
        very_poor_count=very_poor,
        students_files=students_files,
        teachers_files=teachers_files,
        feedbacks=feedbacks,
        feedback_branch=feedback_branch,
        feedback_semester=feedback_semester,
        feedback_subject_code=feedback_subject_code,
        feedback_teacher=feedback_teacher,
        current_feedback_window=current_feedback_window
    
    )
@app.route('/set_feedback_window', methods=['POST'])
@login_required
def set_feedback_window():
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')

    if not start_time or not end_time:
        flash('Start time and end time are required!', 'danger')
        return redirect(url_for('admin_dashboard', tab='feedback'))

    try:
        start_dt = datetime.strptime(start_time, '%Y-%m-%dT%H:%M')
        end_dt = datetime.strptime(end_time, '%Y-%m-%dT%H:%M')

        if end_dt <= start_dt:
            flash('End time must be after start time!', 'danger')
            return redirect(url_for('admin_dashboard', tab='feedback'))

        # Purani active windows deactivate
        FeedbackWindow.query.update({'is_active': False})

        new_window = FeedbackWindow(
            start_time=start_dt,
            end_time=end_dt,
            is_active=True
        )

        db.session.add(new_window)
        db.session.commit()

        flash('Feedback time window set successfully!', 'success')

    except Exception as e:

        flash(f'Error: {str(e)}', 'danger')

    return redirect(url_for('admin_dashboard', tab='feedback'))
# ========================================
# feedbacks view for admin with filters
# ========================================
@app.route('/admin/feedbacks')
@login_required
def admin_feedbacks():
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    branch = request.args.get('branch')
    semester = request.args.get('semester', type=int)
    subject = request.args.get('subject')
    teacher_id = request.args.get('teacher_id', type=int)
    student_name = request.args.get('student_name')

    query = Feedback.query.join(User, Feedback.student_id == User.id)

    # Branch admin ko sirf apni branch ka data dikhe
    if getattr(current_user, 'sub_role', None) == 'branchadmin':
        query = query.filter(User.branch == current_user.branch)
    elif branch:
        query = query.filter(User.branch == branch)

    if semester:
        query = query.filter(Feedback.semester == semester)

    if subject:
        query = query.filter(Feedback.subject.ilike(f"%{subject}%"))

    if teacher_id:
        query = query.filter(Feedback.teacher_id == teacher_id)

    if student_name:
        query = query.filter(User.name.ilike(f"%{student_name}%"))

    feedbacks = query.order_by(Feedback.created_at.desc()).all()

    teachers = User.query.filter_by(role='teacher').all()
    if getattr(current_user, 'sub_role', None) == 'branchadmin':
        teachers = User.query.filter_by(role='teacher', branch=current_user.branch).all()

    return render_template(
        'admin/view_feedbacks.html',
        feedbacks=feedbacks,
        branches=get_all_branches(),
        teachers=teachers,
        selected_branch=branch,
        selected_semester=semester,
        selected_subject=subject,
        selected_teacher_id=teacher_id,
        selected_student_name=student_name
    )

# =============================
# feedbacks export for admin 
# =============================

@app.route('/export_admin_feedbacks')
@login_required
def export_admin_feedbacks():
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    branch = request.args.get('branch')
    semester = request.args.get('semester', type=int)
    subject_code = request.args.get('subject_code')
    teacher = request.args.get('teacher', type=int)

    query = Feedback.query.join(User, Feedback.student_id == User.id)

    if branch:
        query = query.filter(User.branch == branch)

    if semester:
        query = query.filter(Feedback.semester == semester)

    if subject_code:
        query = query.filter(Feedback.subject_code == subject_code)

    if teacher:
        query = query.filter(Feedback.teacher_id == teacher)

    feedbacks = query.order_by(Feedback.created_at.desc()).all()

    data = []
    for f in feedbacks:
        data.append({
            "Student Name": f.student.name or f.student.username,
            "Roll No": f.student.roll_number,
            "Teacher Name": f.teacher.name or f.teacher.username,
            "Subject": f.subject,
            "Subject Code": f.subject_code if f.subject_code else "N/A",
            "Branch": f.student.branch,
            "Semester": f.semester,
            "Academic Year": f.academic_year or "N/A",
            "Timestamp": f.created_at.strftime('%Y-%m-%d %H:%M') if f.created_at else "N/A",
            "Q1": f.q1,
            "Q2": f.q2,
            "Q3": f.q3,
            "Q4": f.q4,
            "Q5": f.q5,
            "Q6": f.q6,
            "Q7": f.q7,
            "Q8": f.q8,
            "Q9": f.q9,
            "Q10": f.q10,
            "Q11": f.q11,
            "Q12": f.q12,
            "Total": f.total,
            "Average": round(f.total / 12, 2) if f.total else 0
        })

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Feedbacks')

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='admin_feedbacks.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ========================
# SUPERADMIN DASHBOARD
# ========================
@app.route('/superadmin_dashboard')
@login_required
def superadmin_dashboard():
    role = getattr(current_user, 'role', '').lower()  # safe way
    if role != 'superadmin':
        flash("Access denied!", "danger")
        return redirect(url_for('home'))  # ya koi public page
       
    # Agar role correct hai, files fetch karo
    students_files = UploadedFile.query.filter_by(file_type='student').all()
    teachers_files = UploadedFile.query.filter_by(file_type='teacher').all()
    admins_files = UploadedFile.query.filter_by(file_type='admin').all()

    return render_template(
        'admin/superadmin_dashboard.html',
        students_files=students_files,
        teachers_files=teachers_files,
        admins_files=admins_files
    )
@app.route('/add_admin', methods=['POST'])
@login_required
def add_admin():
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    username = request.form['username']
    password = request.form['password']
    sub_role = request.form['sub_role']
    branch = request.form.get('branch') if sub_role=='branchadmin' else None
    hashed_pw = generate_password_hash(password)
    new_admin = User(username=username, password=hashed_pw, role='admin', sub_role=sub_role, branch=branch)
    db.session.add(new_admin)
    db.session.commit()
    flash('New admin added successfully!', 'success')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/delete_admin/<int:admin_id>')
@login_required
def delete_admin(admin_id):
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    admin = User.query.get(admin_id)
    if admin and admin.sub_role != 'superadmin':
        db.session.delete(admin)
        db.session.commit()
        flash('Admin deleted successfully!', 'success')
    else:
        flash('Cannot delete superadmin!', 'danger')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/change_admin_password/<int:admin_id>', methods=['POST'])
@login_required
def change_admin_password(admin_id):
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    new_password = request.form['new_password']
    admin = User.query.get(admin_id)
    admin.password = generate_password_hash(new_password)
    db.session.commit()
    flash('Password changed successfully!', 'success')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/change_admin_branch/<int:admin_id>', methods=['POST'])
@login_required
def change_admin_branch(admin_id):
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    new_branch = request.form['new_branch']
    admin = User.query.get(admin_id)
    admin.branch = new_branch
    db.session.commit()
    flash('Branch updated successfully!', 'success')
    return redirect(url_for('superadmin_dashboard'))

# ========================
# EXCEL UPLOAD / DELETE
# ========================
@app.route('/upload_students', methods=['POST'])
@login_required
def upload_students():
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    if 'file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(url_for('admin_dashboard'))

    file = request.files['file']
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        upload_folder = os.path.join(os.getcwd(), 'uploads')
        os.makedirs(upload_folder, exist_ok=True)

        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)

        uploaded_file = UploadedFile(
            filename=filename,
            file_type='student',
            uploaded_by=current_user.id
        )
        db.session.add(uploaded_file)
        db.session.commit()
        flash('Student data synced successfully!', 'success')

        sync_students_from_excel(file_path, uploaded_file.id)

        flash('Students imported/updated successfully!', 'success')

    except Exception as e:
        db.session.rollback()
        print("UPLOAD STUDENTS ERROR:", e)
        import traceback
        traceback.print_exc()

        flash(f'Error: {str(e)}', 'danger')

    return redirect(url_for('admin_dashboard'))

@app.route('/upload_teachers', methods=['POST'])
@login_required
def upload_teachers():
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    if 'file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(url_for('admin_dashboard'))

    file = request.files['file']
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        upload_folder = os.path.join(os.getcwd(), 'uploads')
        os.makedirs(upload_folder, exist_ok=True)

        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)

        print("✅ FILE SAVED:", file_path)

        uploaded_file = UploadedFile(
            filename=filename,
            file_type='teacher',
            uploaded_by=current_user.id
        )
        db.session.add(uploaded_file)
        db.session.commit()

        # 🔥 IMPORTANT: sync call
        sync_teachers_from_excel(file_path, uploaded_file.id)

        print("✅ SYNC COMPLETED")

        flash('Teachers uploaded & synced successfully!', 'success')

    except Exception as e:
        db.session.rollback()
        print("❌ ERROR IN TEACHER UPLOAD:", e)
        import traceback
        traceback.print_exc()

        flash(f'Error: {str(e)}', 'danger')

    return redirect(url_for('admin_dashboard'))

@app.route('/upload_admins', methods=['POST'])
@login_required
def upload_admins():
    if current_user.role != 'admin' or getattr(current_user, 'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))

    if 'file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(url_for('superadmin_dashboard'))

    file = request.files['file']
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('superadmin_dashboard'))

    try:
        upload_folder = os.path.join(os.getcwd(), 'uploads')
        os.makedirs(upload_folder, exist_ok=True)

        filename = file.filename
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)

        uploaded_file = UploadedFile(filename=filename, file_type='admin', uploaded_by=current_user.id)
        db.session.add(uploaded_file)
        db.session.commit()

        wb = load_workbook(file_path)
        ws = wb.active

        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]         
        header_index = {header: i for i, header in enumerate(headers)}

        imported = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            username_idx = header_index.get('username')
            password_idx = header_index.get('password')
            sub_role_idx = header_index.get('sub_role')
            branch_idx = header_index.get('branch')
            name_idx = header_index.get('name')
            email_idx = header_index.get('e-mail')

            username = str(row[username_idx]).strip() if username_idx is not None and row[username_idx] else None
            if not username:
                continue

            if User.query.filter_by(username=username).first():
                skipped += 1
                continue

            email_val = str(row[email_idx]).strip() if email_idx is not None and row[email_idx] and str(row[email_idx]).strip() != "" else None

            user = User(
                username=username,
                password=generate_password_hash(
                    str(row[password_idx]).strip() if password_idx is not None and row[password_idx] else 'admin@123'
                ),
                role='admin',
                sub_role=str(row[sub_role_idx]).strip() if sub_role_idx is not None and row[sub_role_idx] else 'branchadmin',
                branch=str(row[branch_idx]).strip() if branch_idx is not None and row[branch_idx] else None,
                name=str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else '',
                email=email_val,
                imported_file_id=uploaded_file.id
            )

            db.session.add(user)
            imported += 1

        db.session.commit()
        flash(f'Successfully imported {imported} admins! ({skipped} skipped due to duplicate username)', 'success')

    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')

    return redirect(url_for('superadmin_dashboard'))

@app.route('/sync_uploaded_file/<int:file_id>')
@login_required
def sync_uploaded_file(file_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    uploaded_file = UploadedFile.query.get_or_404(file_id)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)

    if not os.path.exists(file_path):
        flash('Physical file missing!', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # Headers ko lowercase aur clean karne ke liye logic
        raw_headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        
        if uploaded_file.file_type == 'student':
            # Student Columns Mapping
            idx_map = {
                'username': flexible_header_match(raw_headers, ['username', 'id']),
                'password': flexible_header_match(raw_headers, ['password', 'pw']),
                'name': flexible_header_match(raw_headers, ['name', 'studentname']),
                'email': flexible_header_match(raw_headers, ['email', 'emailid', 'e-mail']),
                'roll_number': flexible_header_match(raw_headers, ['rollnumber', 'rollno']),
                'branch': flexible_header_match(raw_headers, ['branch', 'dept']),
                'semester': flexible_header_match(raw_headers, ['semester', 'sem']),
                'academic_year': flexible_header_match(raw_headers, ['academicyear', 'year'])
            }

            for row in ws.iter_rows(min_row=2, values_only=True):
                uname_val = str(row[idx_map['username']]).strip().replace('.0', '') if idx_map['username'] is not None and row[idx_map['username']] else None
                if not uname_val:
                    continue

                # Existing student check karein (Username ya Roll number se)
                roll_val = str(row[idx_map['roll_number']]).strip().replace('.0', '') if idx_map['roll_number'] is not None and row[idx_map['roll_number']] else None
                
                user = User.query.filter((User.username == uname_val) | (User.roll_number == roll_val)).filter_by(role='student').first()
                
                pwd_val = str(row[idx_map['password']]).strip() if idx_map['password'] is not None and row[idx_map['password']] else '123456'
                
                if not user:
                    user = User(username=uname_val, role='student')
                    db.session.add(user)

                user.password = generate_password_hash(pwd_val)
                user.name = str(row[idx_map['name']]).strip() if idx_map['name'] is not None and row[idx_map['name']] else uname_val
                user.email = str(row[idx_map['email']]).strip() if idx_map['email'] is not None and row[idx_map['email']] else None
                user.roll_number = roll_val
                user.branch = str(row[idx_map['branch']]).strip() if idx_map['branch'] is not None and row[idx_map['branch']] else ''
                user.academic_year = str(row[idx_map['academic_year']]).strip() if idx_map['academic_year'] is not None and row[idx_map['academic_year']] else ''
                user.imported_file_id = uploaded_file.id
                
                # Semester numeric extract karein
                if idx_map['semester'] is not None and row[idx_map['semester']]:
                    try:
                        sem_str = str(row[idx_map['semester']]).strip()
                        user.semester = int(''.join(filter(str.isdigit, sem_str)))
                    except:
                        user.semester = None

        elif uploaded_file.file_type == 'teacher':
            # Teachers ke liye purani mapping hatayein taaki duplicates na ho
            TeacherSubject.query.filter(TeacherSubject.teacher_id.in_(
                db.session.query(User.id).filter_by(role='teacher')
            )).delete(synchronize_session=False)
            
            # Naya Teacher Mapping Logic (Jo Excel columns aapne dikhaye hain)
            sync_teachers_from_excel(file_path, uploaded_file.id)

        db.session.commit()
        flash('File and Database synced successfully!', 'success')

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        flash(f'Sync Error: {str(e)}', 'danger')

    return redirect(url_for('admin_dashboard'))
@app.route('/sync_to_db/<int:file_id>', methods=['POST'])
@login_required
def sync_to_db(file_id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    uploaded_file = UploadedFile.query.get_or_404(file_id)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)

    if not os.path.exists(file_path):
        return jsonify({'success': False, 'message': 'Physical file missing on server!'}), 404

    try:
        # File type ke basis par sahi function call karein
        if uploaded_file.file_type == 'student':
            sync_students_from_excel(file_path, uploaded_file.id)
        elif uploaded_file.file_type == 'teacher':
            # Purani mapping saaf karein taaki duplicate na ho
            TeacherSubject.query.filter(TeacherSubject.teacher_id.in_(
                db.session.query(User.id).filter_by(role='teacher')
            )).delete(synchronize_session=False)
            
            sync_teachers_from_excel(file_path, uploaded_file.id)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Database synced successfully! Now redirecting...'})
    
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Sync failed: {str(e)}'}), 500
@app.route('/delete_file/<int:file_id>')
@login_required
def delete_file(file_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    file = UploadedFile.query.get(file_id)

    if not file:
        flash("File not found!", "danger")
        return redirect(url_for('admin_dashboard'))

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)

    try:
        # Related users
        users = User.query.filter_by(imported_file_id=file.id).all()

        for user in users:
            # Student ke diye hue feedback
            Feedback.query.filter_by(student_id=user.id).delete(synchronize_session=False)

            # Teacher ko mile hue feedback
            Feedback.query.filter_by(teacher_id=user.id).delete(synchronize_session=False)

            # Teacher-subject mappings
            TeacherSubject.query.filter_by(teacher_id=user.id).delete(synchronize_session=False)

            # User delete
            db.session.delete(user)

        # File record delete from DB
        db.session.delete(file)
        db.session.commit()

        # Physical file delete from uploads folder
        if os.path.exists(file_path):
            os.remove(file_path)

        flash("File, related users, mappings and uploaded file deleted successfully!", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting file: {str(e)}", "danger")

    return redirect(url_for('admin_dashboard'))
# ========================
# VIEW / EDIT / DELETE STUDENTS & TEACHERS
# ========================
@app.route('/view_students')
@login_required
def view_students():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    students = User.query.filter_by(role='student', branch=current_user.branch).all() \
               if getattr(current_user, 'sub_role', None)=='branchadmin' else User.query.filter_by(role='student').all()
    return render_template('admin/view_students.html', users=students)

@app.route('/view_teachers')
@login_required
def view_teachers():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    teachers = User.query.filter_by(role='teacher').all()
    return render_template('admin/view_teachers.html', users=teachers)

@app.route('/edit_user/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    user = User.query.get(user_id)
    if not user:
        flash("User not found!", "danger")
        return redirect(request.referrer)
    
    # Update fields
    user.name = request.form.get('name', user.name)
    user.email = request.form.get('email', user.email)
    if user.role=='student':
        user.branch = request.form.get('branch', user.branch)
        user.semester = int(request.form.get('semester', user.semester))
        user.academic_year = request.form.get('academic_year', user.academic_year)
        user.roll_number = request.form.get('roll_number', user.roll_number)
    elif user.role=='teacher':
        user.subject = request.form.get('subject', user.subject)
    
    db.session.commit()
    flash(f"{user.role.title()} '{user.username}' updated successfully!", "success")
    return redirect(request.referrer)

@app.route('/delete_user/<int:user_id>')
@login_required
def delete_user(user_id):
    user = User.query.get(user_id)

    if not user:
        flash("User not found!", "danger")
        return redirect(url_for('admin_dashboard'))

    # student ke feedback delete
    Feedback.query.filter_by(student_id=user.id).delete(synchronize_session=False)

    # teacher ke feedback delete
    Feedback.query.filter_by(teacher_id=user.id).delete(synchronize_session=False)

    # teacher subject mapping delete
    TeacherSubject.query.filter_by(teacher_id=user.id).delete(synchronize_session=False)

    db.session.delete(user)
    db.session.commit()

    flash("User deleted permanently!", "success")
    return redirect(url_for('admin_dashboard'))

import pandas as pd  # 👈 ye import add karna hai file ke top me
def sync_students_from_excel(file_path, uploaded_file_id):
    wb = load_workbook(file_path)
    ws = wb.active

    def normalize_header(header):
        if not header:
            return ''
        header = str(header).strip().lower()
        header = header.replace(' ', '_')
        header = header.replace('-', '_')
        header = header.replace('.', '')
        return header

    def clean_value(value):
        if value is None:
            return None
        value = str(value).strip()
        if value == '':
            return None
        return value

    headers = [normalize_header(cell.value) for cell in ws[1]]
    header_index = {header: i for i, header in enumerate(headers)}

    # alternate header names handle karne ke liye
    username_idx = header_index.get('username')
    password_idx = header_index.get('password')
    name_idx = header_index.get('name')
    email_idx = header_index.get('email') or header_index.get('e_mail')
    branch_idx = header_index.get('branch')
    semester_idx = header_index.get('semester')
    roll_idx = header_index.get('roll_no') or header_index.get('roll_number')
    academic_year_idx = header_index.get('academic_year')

    for row in ws.iter_rows(min_row=2, values_only=True):
        username = clean_value(row[username_idx]) if username_idx is not None else None
        password = clean_value(row[password_idx]) if password_idx is not None else None
        name = clean_value(row[name_idx]) if name_idx is not None else None
        email = clean_value(row[email_idx]) if email_idx is not None else None
        roll_number = clean_value(row[roll_idx]) if roll_idx is not None else None
        branch = clean_value(row[branch_idx]) if branch_idx is not None else None
        academic_year = clean_value(row[academic_year_idx]) if academic_year_idx is not None else None

        semester = None
        if semester_idx is not None and row[semester_idx] is not None:
            sem_value = str(row[semester_idx]).strip().lower()
            sem_value = sem_value.replace('th', '').replace('st', '').replace('nd', '').replace('rd', '')
            try:
                semester = int(sem_value)
            except:
                semester = None

        if username:
            username = username.replace('.0', '')
        if roll_number:
            roll_number = roll_number.replace('.0', '')

        if not username:
            continue

        existing_user = None

        if roll_number:
            existing_user = User.query.filter_by(role='student', roll_number=roll_number).first()

        if not existing_user and username:
            existing_user = User.query.filter_by(role='student', username=username).first()

        if existing_user:
            if email:
                another_email_user = User.query.filter(
                    User.role == 'student',
                    User.email == email,
                    User.id != existing_user.id
                ).first()
                if another_email_user:
                    email = None

            existing_user.username = username
            existing_user.name = name
            existing_user.email = email
            existing_user.branch = branch
            existing_user.semester = semester
            existing_user.academic_year = academic_year
            existing_user.roll_number = roll_number
            existing_user.imported_file_id = uploaded_file_id

            if password:
                existing_user.password = generate_password_hash(password)

        else:
            if email:
                already_email = User.query.filter_by(role='student', email=email).first()
                if already_email:
                    email = None

            new_user = User(
                username=username,
                password=generate_password_hash(password) if password else generate_password_hash('123456'),
                name=name,
                email=email,
                branch=branch,
                semester=semester,
                academic_year=academic_year,
                roll_number=roll_number,
                role='student',
                imported_file_id=uploaded_file_id
            )
            db.session.add(new_user)

    db.session.commit()
@app.route('/view_file/<int:file_id>')
@login_required
def view_file(file_id):
    uploaded_file = UploadedFile.query.get_or_404(file_id)

    if current_user.role != 'admin':
        return redirect(url_for('home'))

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)

    if not os.path.exists(file_path):
        flash(f"Physical file missing: {uploaded_file.filename}", "danger")
        return redirect(url_for('admin_dashboard'))

    try:
        df = pd.read_excel(file_path)
        df = df.fillna("")

        data = df.values.tolist()
        columns = list(df.columns)

        return render_template(
            'admin/spreadsheet.html',
            file=uploaded_file,
            file_id=uploaded_file.id,
            data=data,
            columns=columns
        )

    except Exception as e:
        flash(f"Unable to open file: {str(e)}", "danger")
        return redirect(url_for('admin_dashboard'))
def flexible_header_match(header_list, targets):
    """Excel headers mein se keywords dhoondhne ke liye dynamic logic"""
    for i, h in enumerate(header_list):
        if h:
            # Header ko clean karein (spaces, dots, special chars hatayein)
            clean_h = str(h).strip().lower().replace(' ', '').replace('_', '').replace('-', '').replace('.', '')
            if clean_h in targets:
                return i
    return None
@app.route('/download_excel_file/<int:file_id>')
@login_required
def download_excel_file(file_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    uploaded_file = UploadedFile.query.get_or_404(file_id)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)
    
    if os.path.exists(file_path):
        from flask import send_file
        return send_file(file_path, as_attachment=True)
    else:
        flash('File not found on server!', 'danger')
        return redirect(url_for('admin_dashboard'))
    

def sync_teachers_from_excel(file_path, uploaded_file_id):
    wb = load_workbook(file_path)
    ws = wb.active
    raw_headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    print("🚀 SYNC FUNCTION STARTED")
    print("HEADERS:", raw_headers)
    
    # Aapki image ke headers se exact match
    idx_map = {
        'username': flexible_header_match(raw_headers, ['username']),
        'email': flexible_header_match(raw_headers, ['e-mail', 'email']),
        'subject': flexible_header_match(raw_headers, ['subject']),
        'password': flexible_header_match(raw_headers, ['password']),
        'sub_code': flexible_header_match(raw_headers, ['subjectcode']),
        'semester': flexible_header_match(raw_headers, ['semester']),
        'branch': flexible_header_match(raw_headers, ['branch'])
    }

    print("INDEX MAP:", idx_map)

    for row in ws.iter_rows(min_row=2, values_only=True):
        print("ROW:", row)

        email_idx = idx_map['email']
        if email_idx is None or row[email_idx] is None: 
            continue
        
        email = str(row[email_idx]).strip()
        
        # 1. Teacher User Account Handle Karein
        teacher = User.query.filter_by(email=email, role='teacher').first()
        raw_pwd = str(row[idx_map['password']]).strip() if idx_map['password'] is not None and row[idx_map['password']] else 'teacher@123'
        
        if not teacher:
            raw_name = str(row[idx_map['username']]).strip() if idx_map['username'] is not None else email.split('@')[0]
            # Username se dots aur spaces hatayein login ke liye
            clean_uname = str(row[idx_map['username']]).strip()
            
            teacher = User(
                username=clean_uname, 
                email=email, 
                role='teacher', 
                password=generate_password_hash(raw_pwd),
                name=raw_name
            )
            db.session.add(teacher)
            db.session.flush() 
        else:
            teacher.password = generate_password_hash(raw_pwd)

        teacher.imported_file_id = uploaded_file_id

        # 2. Subject aur Mapping Handle Karein
        branch_raw = str(row[idx_map['branch']]).strip() if idx_map['branch'] is not None else ""
        sem_raw = str(row[idx_map['semester']]).strip() if idx_map['semester'] is not None else ""
        
        # Semester extract karein (2, 4, 6 etc.)
        sem_num = None
        if sem_raw:
            digits = ''.join(filter(str.isdigit, sem_raw))
            if digits: sem_num = int(digits)

        if branch_raw and sem_num:
            # Normalize branch (e.g., 'CSE(AI&DS)' -> 'cseais')
            branch_value = str(row[idx_map['branch']]).strip()
            sub_name = str(row[idx_map['subject']]).strip() if idx_map['subject'] is not None else ""
            sub_code = str(row[idx_map['sub_code']]).strip() if idx_map['sub_code'] is not None else ""

            # Check mapping to avoid duplicates
            existing_map = TeacherSubject.query.filter_by(
                teacher_id=teacher.id, 
                subject=sub_name, 
                branch=branch_value, 
                semester=sem_num
            ).first()

            if not existing_map:
                new_map = TeacherSubject(
                    teacher_id=teacher.id,
                    subject=sub_name,
                    subject_code=sub_code,
                    branch=branch_value,
                    semester=sem_num
                )
                db.session.add(new_map)

    db.session.commit()


#@app.route('/select_details')
#@login_required
#def select_details():
 #   if current_user.role != 'student':
  #      return redirect(url_for('home'))

   # return render_template('student/select.html', branches=get_all_branches())


@app.route('/set_selection', methods=['POST'])
@login_required
def set_selection():
    branch = request.form.get('branch')
    semester = request.form.get('semester')

    #if not branch or not semester:
     #   flash("Select both branch and semester!", "danger")
      #  return redirect(url_for('select_details'))

    semester = int(semester)

    # ✅ CHECK: kya is branch + semester ke liye subjects exist karte hain?
    exists = [
        t for t in TeacherSubject.query.filter_by(semester=semester).all()
        if normalize_branch(t.branch) == normalize_branch(branch)
    ]

    if not exists:
        flash("Invalid selection!", "danger")
        return redirect(url_for('select_details'))

    # ✅ VALID hai tabhi session me save karo
    session['branch'] = branch  
    session['semester'] = semester

    return redirect(url_for('student_dashboard'))


# ========================
# INITIALIZE DATABASE
# ========================
def create_default_admin():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(username='superadmin').first():
            superadmin = User(username='superadmin', password=generate_password_hash('admin123'),
                              email='superadmin@college.com', role='admin', sub_role='superadmin',
                              name='Super Admin', branch=None)
            db.session.add(superadmin)
        for branch in get_all_branches():
            uname = branch.lower().replace('&','').replace('-','').replace(' ','')+'admin'
            if not User.query.filter_by(username=uname).first():
                admin = User(username=uname, password=generate_password_hash('admin123'),
                             email=f'{uname}@college.com', role='admin', sub_role='branchadmin',
                             name=f'{branch} Admin', branch=branch)
                db.session.add(admin)
        db.session.commit()
        print("✅ Database initialized successfully!")
        print("\n📋 Default Login Credentials:")
        print("-"*50)
        print("| Role          | Username          | Password |")
        print("-"*50)
        print("| Super Admin   | superadmin        | admin123 |")
    
        print("-"*50)

if __name__ == '__main__':
    create_default_admin()
    app.run(debug=True, use_reloader=False)




